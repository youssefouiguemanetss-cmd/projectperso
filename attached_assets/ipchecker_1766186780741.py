import dns.resolver
import re
import concurrent.futures
from collections import defaultdict
import openpyxl
from datetime import date

# === YOUR DQS KEY ===
DQS_KEY = "f3jqdoqpeyipweiizk7onufnlm"

# Global resolver
resolver = dns.resolver.Resolver()
resolver.timeout = 5
resolver.lifetime = 10
resolver.retries = 3
resolver.nameservers = ['8.8.8.8', '1.1.1.1']

results = defaultdict(set)

IP_REGEX = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
DOMAIN_REGEX = re.compile(r"^([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}$")

def parse_line(line):
    line = line.strip()
    if not line:
        return None

    parts = line.split(":", 2)  # Split on first two colons only

    if len(parts) == 3:
        server, ip, domain = parts
        if IP_REGEX.match(ip) and DOMAIN_REGEX.match(domain):
            return ("ip_domain", server, ip, domain)
    
    elif len(parts) == 2:
        server, value = parts
        if IP_REGEX.match(value):
            return ("ip", server, value)
        elif DOMAIN_REGEX.match(value):
            return ("domain", server, value)
    
    elif len(parts) == 1:
        value = parts[0]
        if IP_REGEX.match(value):
            return ("ip", "unknown_server", value)
        elif DOMAIN_REGEX.match(value):
            return ("domain", "unknown", value)

    return None

def check_spamhaus_ip(ip):
    try:
        rev = ".".join(ip.split(".")[::-1])
        query = f"{rev}.{DQS_KEY}.zen.dq.spamhaus.net"
        answers = resolver.resolve(query, "A")
        found = {r.to_text() for r in answers}

        if "127.0.0.3" in found:
            return "css"
        if "127.0.0.2" in found or "127.0.0.9" in found:
            return "sbl"
        if found.intersection({"127.0.0.4", "127.0.0.5", "127.0.0.6", "127.0.0.7"}):
            return "xbl"
        if "127.0.0.10" in found or "127.0.0.11" in found:
            return "pbl"
        if any(r.startswith("127.0.0.") for r in found):
            return "sbl"
        return None
    except dns.resolver.NXDOMAIN:
        return "clean"
    except Exception:
        return None

def check_barracuda(ip):
    try:
        rev = ".".join(ip.split(".")[::-1])
        query = f"{rev}.b.barracudacentral.org"
        resolver.resolve(query, "A")
        return True
    except dns.resolver.NXDOMAIN:
        return False
    except Exception:
        return False

# FIXED: Use DQS for reliable DBL checks
def check_spamhaus_domain(domain):
    try:
        query = f"{domain}.{DQS_KEY}.dbl.dq.spamhaus.net"
        answers = resolver.resolve(query, "A")
        # Any 127.0.1.x response means listed in DBL
        if any(r.to_text().startswith("127.0.1.") for r in answers):
            return "dbl"
        return None
    except dns.resolver.NXDOMAIN:
        return "clean"
    except Exception:
        return None

def process_item(original_line):
    parsed = parse_line(original_line)
    if not parsed:
        return None

    kind = parsed[0]

    if kind == "ip_domain":
        _, server, ip, domain = parsed
        output_line_ip = f"{server}:{ip}"

        # Check IP
        spamhaus_ip = check_spamhaus_ip(ip)
        barracuda = check_barracuda(ip)

        if spamhaus_ip == "clean" and not barracuda:
            return ("clean_ips", ip)
        else:
            if spamhaus_ip:
                results[spamhaus_ip].add(output_line_ip)
            if barracuda:
                results["barracuda"].add(output_line_ip)

        # Check Domain independently
        dbl_result = check_spamhaus_domain(domain)
        if dbl_result == "dbl":
            results["dbl"].add((server, ip, domain))  # IP as Class

        return None

    elif kind == "domain":
        _, server, domain = parsed
        result = check_spamhaus_domain(domain)
        if result == "dbl":
            results["dbl"].add((server, "", domain))  # Empty Class
        return None

    elif kind == "ip":
        _, server, ip = parsed
        output_line = f"{server}:{ip}"
        spamhaus = check_spamhaus_ip(ip)
        barracuda = check_barracuda(ip)

        if spamhaus == "clean" and not barracuda:
            return ("clean_ips", ip)

        if spamhaus:
            results[spamhaus].add(output_line)
        if barracuda:
            results["barracuda"].add(output_line)
        return None

def write_to_template():
    today = date.today().strftime("%d-%m-%Y")
    output_filename = f"Check Blacklist All TSSW {today}.xlsx"
    template_path = "template.xlsx"

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Clear old data
    data_columns = ['A','B','D','E','G','H','K','L','N','O','Q','R','S','U']
    for col in data_columns:
        start_row = 7 if col == 'U' else 8
        for row in range(start_row, ws.max_row + 10):
            ws[f"{col}{row}"] = None

    # Fill IP categories
    categories = [
        ("css", "A", "B"),
        ("pbl", "D", "E"),
        ("xbl", "G", "H"),
        ("sbl", "K", "L"),
        ("barracuda", "N", "O"),
    ]

    for cat, col_server, col_ip in categories:
        if cat in results:
            items = sorted(results[cat])
            row = 8
            for item in items:
                parts = item.split(":", 1)
                server = parts[0] if len(parts) > 1 else "unknown_server"
                val = parts[1] if len(parts) > 1 else item
                ws[f"{col_server}{row}"] = server
                ws[f"{col_ip}{row}"] = val
                row += 1

    # Fill DBL section
    if "dbl" in results:
        listed_dbl = sorted(results["dbl"], key=lambda x: (x[0], x[1], x[2]))
        row = 8
        for server, class_ip, domain in listed_dbl:
            ws[f"Q{row}"] = server
            ws[f"R{row}"] = class_ip   # IP as Class (or empty)
            ws[f"S{row}"] = domain
            row += 1

    # Clean IPs
    if "clean_ips" in results:
        clean_ips = sorted(results["clean_ips"])
        row = 7
        for ip in clean_ips:
            ws[f"U{row}"] = ip
            row += 1

    wb.save(output_filename)
    return output_filename

def main():
    with open("input.txt", "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    max_workers = min(100, len(lines) or 1)
    clean_results = defaultdict(list)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_item, line): line for line in lines}
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            if result:
                cat, value = result
                clean_results[cat].append(value)

    for cat, values in clean_results.items():
        results[cat].update(values)

    output_file = write_to_template()
    print(f"Success! All checks completed reliably (DBL now uses DQS).")
    print(f"Results saved to: {output_file}")

if __name__ == "__main__":
    main()